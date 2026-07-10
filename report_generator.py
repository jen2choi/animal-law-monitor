"""
주간 입법 동향 리포트 생성기 v3 - ALLBILLV2 필드 반영
6개 시트: 대시보드 / 계류중 / 처리완료 / 이번주변동 / 전체발의안 / 트렌드통계
"""
import logging
from datetime import datetime, timedelta
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from database import get_conn
from config import REPORT_DIR

logger = logging.getLogger(__name__)
REPORT_DIR_PATH = Path(REPORT_DIR)

# 색상
GREEN_DARK  = "1B4332"
GREEN_MID   = "2D6A4F"
GREEN_SOFT  = "52B788"
GREEN_LIGHT = "D8F3DC"
AMBER       = "E9A825"
AMBER_LIGHT = "FFF3E0"
RED_SOFT    = "E63946"
RED_LIGHT   = "FDECEA"
BLUE_SOFT   = "457B9D"
BLUE_LIGHT  = "E8F4FD"
GRAY_LIGHT  = "F5F5F5"
GRAY_MID    = "CCCCCC"
WHITE       = "FFFFFF"
BLACK       = "1A1A1A"


# ── 스타일 헬퍼 ──

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _font(bold=False, size=11, color=BLACK, name="맑은 고딕"):
    return Font(bold=bold, size=size, color=color, name=name)

def _border():
    s = Side(style="thin", color=GRAY_MID)
    return Border(left=s, right=s, top=s, bottom=s)

def _align(h="center", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def header_row(ws, row, cols, bg=GREEN_MID, fg=WHITE):
    for c, val in enumerate(cols, 1):
        cell = ws.cell(row=row, column=c, value=val)
        cell.fill = _fill(bg); cell.font = _font(bold=True, color=fg)
        cell.border = _border(); cell.alignment = _align()

def data_row(ws, row, values, bg=None, shade=False):
    fill = bg or (GRAY_LIGHT if shade else WHITE)
    for c, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=c, value=val)
        cell.fill = _fill(fill); cell.font = _font()
        cell.border = _border(); cell.alignment = _align(h="left", wrap=True)

def section_title(ws, row, text, max_col, bg=GREEN_SOFT):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max_col)
    c = ws.cell(row=row, column=1, value=text)
    c.fill = _fill(bg); c.font = _font(bold=True, size=12, color=WHITE)
    c.alignment = _align(); ws.row_dimensions[row].height = 28


# ── 데이터 쿼리 ──

def get_all_data(start: str, end: str) -> dict:
    with get_conn() as conn:

        pending = conn.execute("""
            SELECT bill_id, bill_no, bill_name, bill_kind, proposer, proposer_kind,
                   propose_dt, propose_sess, committee,
                   committee_proc_result, proc_result, detail_link
            FROM bills
            WHERE proc_result IS NULL OR proc_result = ''
            ORDER BY propose_dt DESC
        """).fetchall()

        completed = conn.execute("""
            SELECT bill_id, bill_no, bill_name, bill_kind, proposer, proposer_kind,
                   propose_dt, committee, committee_proc_result,
                   proc_result, proc_dt, detail_link
            FROM bills
            WHERE proc_result IS NOT NULL AND proc_result != ''
            ORDER BY proc_dt DESC NULLS LAST
        """).fetchall()

        new_bills = conn.execute("""
            SELECT bill_id, bill_no, bill_name, bill_kind, proposer, proposer_kind,
                   propose_dt, committee, proc_result, detail_link
            FROM bills
            WHERE first_seen BETWEEN ? AND ?
            ORDER BY propose_dt DESC
        """, (start, end)).fetchall()

        changed = conn.execute("""
            SELECT h.bill_id, b.bill_no, b.bill_name, b.proposer, h.field_name,
                   h.old_value, h.new_value, h.changed_at, b.detail_link
            FROM bill_history h
            JOIN bills b ON b.bill_id = h.bill_id
            WHERE h.changed_at BETWEEN ? AND ?
            ORDER BY h.changed_at DESC
        """, (start, end)).fetchall()

        all_bills = conn.execute("""
            SELECT bill_id, bill_no, bill_name, bill_kind, proposer, proposer_kind,
                   propose_dt, propose_sess, committee,
                   committee_proc_result, proc_result, proc_dt,
                   first_seen, last_updated, detail_link,
            ai_score, ai_tags
    FROM bills
            ORDER BY propose_dt DESC
        """).fetchall()

        proc_dist = conn.execute("""
            SELECT COALESCE(NULLIF(proc_result,''),'계류중') AS status, COUNT(*) AS cnt
            FROM bills GROUP BY status ORDER BY cnt DESC
        """).fetchall()

        cmmt_proc_dist = conn.execute("""
            SELECT COALESCE(NULLIF(committee_proc_result,''),'미처리') AS status, COUNT(*) AS cnt
            FROM bills GROUP BY status ORDER BY cnt DESC
        """).fetchall()

        committee_dist = conn.execute("""
            SELECT COALESCE(committee,'미지정') AS committee, COUNT(*) AS cnt
            FROM bills GROUP BY committee ORDER BY cnt DESC LIMIT 15
        """).fetchall()

        bill_kind_dist = conn.execute("""
            SELECT COALESCE(bill_kind,'미분류') AS kind, COUNT(*) AS cnt
            FROM bills GROUP BY kind ORDER BY cnt DESC
        """).fetchall()

        proposer_kind_dist = conn.execute("""
            SELECT COALESCE(proposer_kind,'미분류') AS kind, COUNT(*) AS cnt
            FROM bills GROUP BY kind ORDER BY cnt DESC
        """).fetchall()

        yearly = conn.execute("""
            SELECT SUBSTR(propose_dt,1,4) AS year, COUNT(*) AS cnt
            FROM bills WHERE propose_dt IS NOT NULL AND propose_dt != ''
            GROUP BY year ORDER BY year DESC LIMIT 7
        """).fetchall()

        total    = conn.execute("SELECT COUNT(*) FROM bills").fetchone()[0]
        passed   = conn.execute("SELECT COUNT(*) FROM bills WHERE proc_result LIKE '%가결%'").fetchone()[0]
        rejected = conn.execute("SELECT COUNT(*) FROM bills WHERE proc_result LIKE '%부결%' OR proc_result LIKE '%폐기%'").fetchone()[0]

    pending_cnt   = len(pending)
    completed_cnt = len(completed)

    return {
        "pending": [dict(r) for r in pending],
        "completed": [dict(r) for r in completed],
        "new_bills": [dict(r) for r in new_bills],
        "changed": [dict(r) for r in changed],
        "all_bills": [dict(r) for r in all_bills],
        "proc_dist": [dict(r) for r in proc_dist],
        "cmmt_proc_dist": [dict(r) for r in cmmt_proc_dist],
        "committee_dist": [dict(r) for r in committee_dist],
        "bill_kind_dist": [dict(r) for r in bill_kind_dist],
        "proposer_kind_dist": [dict(r) for r in proposer_kind_dist],
        "yearly": [dict(r) for r in yearly],
        "total": total,
        "pending_cnt": pending_cnt,
        "completed_cnt": completed_cnt,
        "passed": passed,
        "rejected": rejected,
    }


# ── 시트 1: 대시보드 ──

def sheet_dashboard(wb, data, start, end):
    ws = wb.active
    ws.title = "📊 대시보드"
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, [22, 18, 22, 18, 22, 18])

    ws.merge_cells("A1:F1")
    ws.row_dimensions[1].height = 50
    c = ws["A1"]
    c.value = "동물보호법 발의안 주간 모니터링 리포트"
    c.fill = _fill(GREEN_DARK); c.font = _font(bold=True, size=18, color=WHITE)
    c.alignment = _align()

    ws.merge_cells("A2:F2")
    ws.row_dimensions[2].height = 24
    c = ws["A2"]
    c.value = f"기준 기간: {start[:10]} ~ {end[:10]}  |  생성일: {datetime.now().strftime('%Y년 %m월 %d일 %H:%M')}"
    c.fill = _fill(GREEN_MID); c.font = _font(size=11, color=WHITE)
    c.alignment = _align()

    ws.row_dimensions[3].height = 16

    cards = [
        ("전체 발의안",   data["total"],         GREEN_MID,  WHITE),
        ("계류 중",       data["pending_cnt"],    AMBER,      WHITE),
        ("처리 완료",     data["completed_cnt"],  BLUE_SOFT,  WHITE),
        ("가결",          data["passed"],         GREEN_SOFT, WHITE),
        ("부결/폐기",     data["rejected"],       RED_SOFT,   WHITE),
        ("이번 주 신규",  len(data["new_bills"]), GREEN_DARK, WHITE),
    ]

    for idx, (label, value, bg, fg) in enumerate(cards, 1):
        ws.row_dimensions[4].height = 22
        ws.row_dimensions[5].height = 36
        ws.row_dimensions[6].height = 22
        ws.cell(row=4, column=idx, value=label).fill = _fill(bg)
        ws.cell(row=4, column=idx).font = _font(size=10, color=fg)
        ws.cell(row=4, column=idx).alignment = _align()
        ws.cell(row=5, column=idx, value=f"{value:,}건").fill = _fill(bg)
        ws.cell(row=5, column=idx).font = _font(bold=True, size=20, color=fg)
        ws.cell(row=5, column=idx).alignment = _align()
        ws.cell(row=6, column=idx).fill = _fill(bg)

    ws.row_dimensions[7].height = 16

    # 계류 중 법안 미리보기
    section_title(ws, 8, "⏳ 계류 중인 법안 현황 (최근 10건)", 6)
    header_row(ws, 9, ["의안번호", "의안명", "의안종류", "대표발의자", "소관위원회", "위원회처리"])
    for i, b in enumerate(data["pending"][:10], 10):
        data_row(ws, i, [
            b["bill_no"], b["bill_name"], b["bill_kind"] or "-",
            b["proposer"] or "-", b["committee"] or "-",
            b["committee_proc_result"] or "미처리"
        ], shade=(i % 2 == 0))

    r = 21
    ws.row_dimensions[r].height = 16
    r += 1

    # 이번 주 변동
    section_title(ws, r, "🔔 이번 주 변동 사항", 6, bg=AMBER)
    r += 1
    header_row(ws, r, ["구분", "의안번호", "의안명", "발의자", "변경 내용", "일시"], bg=AMBER, fg=WHITE)
    r += 1
    field_labels = {
        "proc_result": "본회의 심의결과 변경",
        "committee_proc_result": "위원회 처리결과 변경",
        "committee": "소관위원회 변경",
        "bill_name": "의안명 변경"
    }
    if data["new_bills"]:
        for b in data["new_bills"][:5]:
            data_row(ws, r, ["신규발의", b["bill_no"], b["bill_name"],
                             b["proposer"] or "-", f"신규 발의 ({b['bill_kind'] or '-'})",
                             b["propose_dt"] or "-"], bg=GREEN_LIGHT)
            r += 1
    if data["changed"]:
        for c in data["changed"][:5]:
            data_row(ws, r, ["상태변경", c["bill_id"], c["bill_name"], "-",
                             f"{field_labels.get(c['field_name'], c['field_name'])}: {c['old_value'] or '-'} → {c['new_value'] or '-'}",
                             c["changed_at"][:16]], bg=AMBER_LIGHT)
            r += 1
    if not data["new_bills"] and not data["changed"]:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
        ws.cell(row=r, column=1, value="이번 주 변동 사항이 없습니다.").alignment = _align()
        ws.cell(row=r, column=1).font = _font(color="888888")


# ── 시트 2: 계류 중인 법안 ──

def sheet_pending(wb, data):
    ws = wb.create_sheet("⏳ 계류 중인 법안")
    set_col_widths(ws, [14, 12, 45, 10, 14, 12, 24, 16, 40])
    header_row(ws, 1, ["의안번호", "의안종류", "의안명", "제안자구분", "대표발의자",
                        "발의일", "소관위원회", "위원회처리결과", "링크"], bg=AMBER, fg=WHITE)
    if data["pending"]:
        for i, b in enumerate(data["pending"], 2):
            data_row(ws, i, [
                b["bill_no"], b["bill_kind"] or "-", b["bill_name"],
                b["proposer_kind"] or "-", b["proposer"] or "-",
                b["propose_dt"] or "-", b["committee"] or "-",
                b["committee_proc_result"] or "미처리",
                b["detail_link"] or "-"
            ], shade=(i % 2 == 0))
    else:
        ws.merge_cells("A2:I2")
        ws["A2"].value = "계류 중인 법안이 없습니다."
        ws["A2"].alignment = _align()
        ws["A2"].font = _font(color="888888")


# ── 시트 3: 처리 완료 ──

def sheet_completed(wb, data):
    ws = wb.create_sheet("✅ 처리 완료")
    set_col_widths(ws, [14, 12, 45, 14, 12, 24, 16, 14, 12, 40])
    header_row(ws, 1, ["의안번호", "의안종류", "의안명", "대표발의자", "발의일",
                        "소관위원회", "위원회처리결과", "본회의결과", "처리일", "링크"], bg=BLUE_SOFT, fg=WHITE)
    if data["completed"]:
        for i, b in enumerate(data["completed"], 2):
            if "가결" in (b["proc_result"] or ""):
                bg = GREEN_LIGHT
            elif any(k in (b["proc_result"] or "") for k in ["부결", "폐기"]):
                bg = RED_LIGHT
            else:
                bg = GRAY_LIGHT if i % 2 == 0 else WHITE
            data_row(ws, i, [
                b["bill_no"], b["bill_kind"] or "-", b["bill_name"],
                b["proposer"] or "-", b["propose_dt"] or "-",
                b["committee"] or "-", b["committee_proc_result"] or "-",
                b["proc_result"] or "-", b["proc_dt"] or "-",
                b["detail_link"] or "-"
            ], bg=bg)
    else:
        ws.merge_cells("A2:J2")
        ws["A2"].value = "처리 완료된 법안이 없습니다."
        ws["A2"].alignment = _align()
        ws["A2"].font = _font(color="888888")


# ── 시트 4: 이번 주 변동 ──

def sheet_weekly_changes(wb, data, start, end):
    ws = wb.create_sheet("🔔 이번 주 변동")
    set_col_widths(ws, [14, 12, 45, 14, 12, 24, 18])

    section_title(ws, 1, f"신규 발의안  ({start[:10]} ~ {end[:10]})", 7, bg=GREEN_SOFT)
    header_row(ws, 2, ["의안번호", "의안종류", "의안명", "제안자구분", "대표발의자", "소관위원회", "발의일"])
    r = 3
    if data["new_bills"]:
        for b in data["new_bills"]:
            data_row(ws, r, [
                b["bill_no"], b["bill_kind"] or "-", b["bill_name"],
                b["proposer_kind"] or "-", b["proposer"] or "-",
                b["committee"] or "-", b["propose_dt"] or "-"
            ], bg=GREEN_LIGHT)
            r += 1
    else:
        ws.merge_cells(f"A3:G3")
        ws["A3"].value = "이번 주 신규 발의안이 없습니다."
        ws["A3"].alignment = _align()
        ws["A3"].font = _font(color="888888")
        r = 4

    r += 1
    section_title(ws, r, "상태 변경 법안", 7, bg=AMBER)
    r += 1
    header_row(ws, r, ["의안번호", "의안명", "변경 항목", "이전 값", "현재 값", "변경일시", "링크"], bg=AMBER, fg=WHITE)
    r += 1
    field_labels = {
        "proc_result": "본회의 심의결과",
        "committee_proc_result": "위원회 처리결과",
        "committee": "소관위원회",
        "bill_name": "의안명"
    }
    if data["changed"]:
        for i, c in enumerate(data["changed"]):
            data_row(ws, r, [
                c["bill_id"], c["bill_name"],
                field_labels.get(c["field_name"], c["field_name"]),
                c["old_value"] or "-", c["new_value"] or "-",
                c["changed_at"][:16], c["detail_link"] or "-"
            ], bg=AMBER_LIGHT if i % 2 == 0 else WHITE)
            r += 1
    else:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
        ws.cell(row=r, column=1, value="이번 주 상태 변경 법안이 없습니다.").alignment = _align()
        ws.cell(row=r, column=1).font = _font(color="888888")


# ── 시트 5: 전체 발의안 ──

def sheet_all_bills(wb, data):
    ws = wb.create_sheet("📋 전체 발의안")
    set_col_widths(ws, [14, 12, 45, 10, 14, 12, 8, 24, 16, 14, 12, 40])
    header_row(ws, 1, [
        "의안번호", "의안종류", "의안명", "제안자구분", "대표발의자",
        "발의일", "회기", "소관위원회", "위원회처리결과",
        "본회의결과", "처리일", "링크"
    ])
    for i, b in enumerate(data["all_bills"], 2):
        if "가결" in (b["proc_result"] or ""):
            bg = GREEN_LIGHT
        elif any(k in (b["proc_result"] or "") for k in ["부결", "폐기"]):
            bg = RED_LIGHT
        elif not b["proc_result"]:
            bg = AMBER_LIGHT if i % 2 == 0 else WHITE
        else:
            bg = GRAY_LIGHT if i % 2 == 0 else WHITE
        data_row(ws, i, [
            b["bill_no"], b["bill_kind"] or "-", b["bill_name"],
            b["proposer_kind"] or "-", b["proposer"] or "-",
            b["propose_dt"] or "-", b["propose_sess"] or "-",
            b["committee"] or "-", b["committee_proc_result"] or "-",
            b["proc_result"] or "계류중", b["proc_dt"] or "-",
            b["detail_link"] or "-"
        ], bg=bg)


# ── 시트 6: 트렌드 통계 ──

def sheet_trend(wb, data):
    ws = wb.create_sheet("📈 트렌드 통계")
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, [30, 15, 30, 15])
    row = 1

    def write_section(title, headers, rows_data, bg=GREEN_SOFT):
        nonlocal row
        section_title(ws, row, title, 2, bg=bg)
        row += 1
        header_row(ws, row, headers)
        row += 1
        for i, r in enumerate(rows_data):
            data_row(ws, row, list(r.values())[:2], shade=(i % 2 == 0))
            row += 1
        row += 1

    write_section("본회의 심의결과 분포", ["처리결과", "건수"],
                  [{"s": r["status"], "c": f"{r['cnt']:,}건"} for r in data["proc_dist"]])

    write_section("위원회 처리결과 분포", ["처리결과", "건수"],
                  [{"s": r["status"], "c": f"{r['cnt']:,}건"} for r in data["cmmt_proc_dist"]],
                  bg=BLUE_SOFT)

    write_section("의안종류별 현황", ["의안종류", "건수"],
                  [{"s": r["kind"], "c": f"{r['cnt']:,}건"} for r in data["bill_kind_dist"]],
                  bg=GREEN_MID)

    write_section("제안자구분별 현황", ["제안자구분", "건수"],
                  [{"s": r["kind"], "c": f"{r['cnt']:,}건"} for r in data["proposer_kind_dist"]],
                  bg=AMBER)

    write_section("소관위원회별 현황 (상위 15개)", ["위원회", "건수"],
                  [{"s": r["committee"], "c": f"{r['cnt']:,}건"} for r in data["committee_dist"]])

    write_section("연도별 발의 추이 (최근 7년)", ["연도", "발의 건수"],
                  [{"s": r["year"], "c": f"{r['cnt']:,}건"} for r in data["yearly"]],
                  bg=BLUE_SOFT)

    # 처리율 요약
    total = data["total"] or 1
    section_title(ws, row, "처리율 요약", 2, bg=GREEN_DARK)
    row += 1
    header_row(ws, row, ["항목", "수치"], bg=GREEN_DARK)
    row += 1
    summaries = [
        ("전체 발의안",  f"{data['total']:,}건"),
        ("계류 중",      f"{data['pending_cnt']:,}건 ({data['pending_cnt']/total*100:.1f}%)"),
        ("처리 완료",    f"{data['completed_cnt']:,}건 ({data['completed_cnt']/total*100:.1f}%)"),
        ("가결",         f"{data['passed']:,}건 ({data['passed']/total*100:.1f}%)"),
        ("부결/폐기",    f"{data['rejected']:,}건 ({data['rejected']/total*100:.1f}%)"),
    ]
    for i, (k, v) in enumerate(summaries):
        data_row(ws, row, [k, v], shade=(i % 2 == 0))
        row += 1


# ── 메인 ──

def generate_weekly_report(end_dt=None):
    if end_dt is None:
        end_dt = datetime.now()
    start_dt  = end_dt - timedelta(days=7)
    start_str = start_dt.strftime("%Y-%m-%d 00:00:00")
    end_str   = end_dt.strftime("%Y-%m-%d 23:59:59")

    logger.info("리포트 생성 시작: %s ~ %s", start_str[:10], end_str[:10])
    data = get_all_data(start_str, end_str)

    wb = Workbook()
    sheet_dashboard(wb, data, start_str, end_str)
    sheet_pending(wb, data)
    sheet_completed(wb, data)
    sheet_weekly_changes(wb, data, start_str, end_str)
    sheet_all_bills(wb, data)
    sheet_trend(wb, data)

    REPORT_DIR_PATH.mkdir(parents=True, exist_ok=True)
    date_tag = end_dt.strftime("%Y%m%d_%H%M")
    out_path = REPORT_DIR_PATH / f"동물보호법_주간리포트_{date_tag}.xlsx"

    wb.save(out_path)
    logger.info("리포트 저장 완료: %s", out_path)
    return out_path


def get_report_data(end_dt=None):
    """sheets_uploader에서 사용할 데이터 반환"""
    if end_dt is None:
        end_dt = datetime.now()
    start_dt  = end_dt - timedelta(days=7)
    start_str = start_dt.strftime("%Y-%m-%d 00:00:00")
    end_str   = end_dt.strftime("%Y-%m-%d 23:59:59")
    return get_all_data(start_str, end_str), start_str, end_str


if __name__ == "__main__":
    logging.basicConfig(level=logging.INFO,
                        format="%(asctime)s [%(levelname)s] %(message)s")
    path = generate_weekly_report()
    print(f"\n리포트 생성 완료: {path}")
